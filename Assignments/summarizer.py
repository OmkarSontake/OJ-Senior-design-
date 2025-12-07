import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

FILE = r"C:\Users\dedhi\OneDrive\Desktop\Project\All_Resolved\All_Resolved_Merged.xlsx"
OUT  = r"C:\Users\dedhi\OneDrive\Desktop\Project\All_Resolved\All_Resolved_Merged_WithSummary.xlsx"

def extract_summary(df, sheetname):
    side = "Buy" if "Buy" in sheetname else "Sell"

    # Entry timestamp
    ts_entry = df["ts_event"].iloc[0] if "ts_event" in df else None

    # Resolution logic
    prof = df.get("Profit Filled")
    stop = df.get("Stop Filled")

    if prof is not None and prof.notna().any():
        idx = prof.first_valid_index()
        ts_res = df.loc[idx, "ts_event"]
        res_price = prof.loc[idx]
        flag = "Profit Filled"
    elif stop is not None and stop.notna().any():
        idx = stop.first_valid_index()
        ts_res = df.loc[idx, "ts_event"]
        res_price = stop.loc[idx]
        flag = "Stop Filled"
    else:
        ts_res = "Not Resolved"
        res_price = None
        flag = None

    # Open Price logic
    open_price = None
    if "Buy Open Price" in df:
        vals = df["Buy Open Price"].dropna()
        if len(vals): open_price = vals.iloc[0]

    if open_price is None and "Sell Open Price" in df:
        vals = df["Sell Open Price"].dropna()
        if len(vals): open_price = vals.iloc[0]

    # Profit/Loss
    pl = df["P/L"].dropna().iloc[0] if "P/L" in df and df["P/L"].notna().any() else None

    return {
        "Sheet": sheetname,
        "Buy/Sell": side,
        "TS Entry": ts_entry,
        "TS Resolved": ts_res,
        "Open Price": open_price,
        "Resolution Price": res_price,
        "Stop/Profit Filled": flag,
        "P/L": pl
    }


# -----------------------
# Styling
# -----------------------
bold = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
center = Alignment(horizontal="center")

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# -----------------------
# Load Workbook
# -----------------------
wb = load_workbook(FILE)
sheetnames = wb.sheetnames.copy()  # list all sheets
print(f"Total sheets found: {len(sheetnames)}")

aggregated = []

# -----------------------
# Process each data sheet
# -----------------------
for sheetname in sheetnames:
    if sheetname == "Summary": 
        continue  # skip the simple summary sheet created earlier

    ws = wb[sheetname]

    # load full sheet data (no skiprows anymore)
    df = pd.read_excel(FILE, sheet_name=sheetname)

    if df.empty:
        continue

    summary = extract_summary(df, sheetname)
    aggregated.append(summary)

    # Insert summary block at the top
    ws.insert_rows(1, amount=12)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(row=1, column=1, value="SUMMARY TABLE").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    row = 3
    for key, value in summary.items():
        if key == "Sheet":
            continue

        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)

        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border

        row += 1

# -----------------------
# Create Aggregated Summary Sheet
# -----------------------
summary_df = pd.DataFrame(aggregated)

wb.create_sheet("AGGREGATED SUMMARY", 0)
ws0 = wb["AGGREGATED SUMMARY"]

# write header
for col_i, col_name in enumerate(summary_df.columns, start=1):
    cell = ws0.cell(row=1, column=col_i, value=col_name)
    cell.font = bold
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center

# write rows
for row_i, row_data in enumerate(summary_df.values, start=2):
    for col_i, value in enumerate(row_data, start=1):
        cell = ws0.cell(row=row_i, column=col_i, value=value)
        cell.border = border

# auto column width
for col in ws0.columns:
    max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
    ws0.column_dimensions[col[0].column_letter].width = max_len + 2

# -----------------------
# Save file
# -----------------------
wb.save(OUT)

print("\n✔ All sheet summaries added successfully.")
print("✔ Aggregated summary sheet created at the TOP.")
print("Saved to:", OUT)
